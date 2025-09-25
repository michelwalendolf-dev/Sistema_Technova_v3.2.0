import customtkinter as ctk 
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageDraw
import json
import os
import re
from datetime import datetime
import sys
import subprocess
import csv
import pandas as pd
import math
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import time

ARQUIVO_USUARIOS = "usuarios.json"
ARQUIVO_PESSOAS = {
    "cliente": "clientes.json",
    "funcionario": "funcionarios.json",
    "fornecedor": "fornecedores.json"
}
SEM_IMAGEM = "fotos/sem_imagem.png"

def tratar_valor_vazio(valor):
    if valor is None:
        return "N/A"
    try:
        if isinstance(valor, float) and math.isnan(valor):
            return "N/A"
    except TypeError:
        pass
    if isinstance(valor, str) and (not valor.strip() or valor.strip().lower() == 'nan'):
        return "N/A"
    return valor

def carregar_dados_banco(nome_arquivo):
    if not os.path.exists(nome_arquivo):
        with open(nome_arquivo, "w") as f:
            json.dump([], f)
    with open(nome_arquivo, "r") as f:
        dados = json.load(f)
        for item in dados:
            for chave, valor in item.items():
                item[chave] = tratar_valor_vazio(valor)
        return dados

def salvar_dados_banco(nome_arquivo, data):
    dados_tratados = []
    for item in data:
        item_tratado = {}
        for chave, valor in item.items():
            item_tratado[chave] = tratar_valor_vazio(valor)
        dados_tratados.append(item_tratado)
    
    with open(nome_arquivo, "w") as f:
        json.dump(dados_tratados, f, indent=4)

def validar_email(email):
    return re.match(r"[^@]+@[^@]+\.[^@]+", email)

def validar_cep(cep):
    return re.match(r"^\d{5}-?\d{3}$", cep)

def validar_cpf(cpf):
    return re.match(r"^\d{3}\.?\d{3}\.?\d{3}-?\d{2}$", cpf)

def validar_cnpj(cnpj):
    return re.match(r"^\d{2}\.?\d{3}\.?\d{3}/\d{4}-\d{2}$", cnpj)

def calcular_idade(data_nasc):
    try:
        nasc = datetime.strptime(data_nasc, "%d/%m/%Y")
        hoje = datetime.today()
        idade = hoje.year - nasc.year - ((hoje.month, hoje.day) < (nasc.month, nasc.day))
        return idade
    except:
        return None

def buscar_pessoa_por_documento(documento):
    for tipo, arquivo in ARQUIVO_PESSOAS.items():
        pessoas = carregar_dados_banco(arquivo)
        for pessoa in pessoas:
            if pessoa.get("cpf") == documento or pessoa.get("cnpj") == documento:
                return pessoa, tipo
    return None, None

class TelaLogin(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Login")
        self.configure(fg_color="#202020")
        self.usuario_var = ctk.StringVar()
        self.senha_var = ctk.StringVar()
        self.captcha_var = ctk.BooleanVar(value=False)
        self.erros = {}
        self.captcha_verified = False
        self.captcha_error = False

        self.verificar_captcha = self._verificar_captcha
        self.animar_verificacao = self._animar_verificacao
        self.login = self._login
        self.perguntar_sistema = self._perguntar_sistema
        self.abrir_registro = self._abrir_registro
        self.abrir_principal = self._abrir_principal
        
        self.build_login()
        self.after_idle(lambda: self.state("zoomed"))
        self.iconbitmap("icones//logo.ico")

    def _verificar_captcha(self):
        if not self.captcha_verified:
            try:
                script_dir = os.path.dirname(os.path.abspath(__file__))
                captcha_script = os.path.join(script_dir, "COD_CAPTCHA.py")
                
                if not os.path.exists(captcha_script):
                    raise FileNotFoundError(f"Arquivo CAPTCHA não encontrado: {captcha_script}")
                
                result = subprocess.run(
                    [sys.executable, captcha_script],
                    capture_output=True,
                    text=True,
                    timeout=120  
                )
                
                if result.returncode == 0:
                    self.captcha_verified = True
                    self.captcha_error = False
                    self.animar_verificacao()
                    self.captcha_var.set(True)
                    self.msg_label.configure(text="✓ verificação concluída!", text_color="green")
                else:
                    self.captcha_error = True
                    self.captcha_var.set(False)
                    error_msg = result.stderr.strip() if result.stderr else "CAPTCHA não concluído"
                    self.msg_label.configure(text="CAPTCHA não verificado", text_color="#e74c3c")
                    self.captcha_checkbox.configure(
                        fg_color="#ffcccc",
                        hover_color="#ffcccc"
                    )
                    
            except FileNotFoundError as e:
                messagebox.showerror("Erro", f"Arquivo CAPTCHA não encontrado: {str(e)}")
            except subprocess.TimeoutExpired:
                self.msg_label.configure(text="Tempo excedido no CAPTCHA", text_color="#e74c3c")

    def _animar_verificacao(self):
        self.captcha_checkbox.configure(
            fg_color="#2ecc71", 
            hover_color="Transparent",
            text="Não sou um robô", 
            font=("Roboto", 18),  
            width=28,  
            height=28,
            state="disabled",
            text_color_disabled="black"   
        )

    def _login(self):
        usuario = self.usuario_var.get()
        senha = self.senha_var.get()
        
        self.erros = {}
        self.usuario_asterisk.configure(text="")
        self.senha_asterisk.configure(text="")
        self.usuario_entry.configure(border_color="#ccc")
        self.senha_entry.configure(border_color="#ccc")
        
        campos_vazios = []
        if not usuario.strip():
            campos_vazios.append("usuário")
            self.usuario_entry.configure(border_color="#ffcccc")
            self.usuario_asterisk.configure(text="×")
        if not senha.strip():
            campos_vazios.append("senha")
            self.senha_entry.configure(border_color="#ffcccc")
            self.senha_asterisk.configure(text="×")
        
        if not self.captcha_verified:
            if campos_vazios:
                mensagem_erro = "Preencha os campos de " + ", ".join(campos_vazios) + " e complete o CAPTCHA"
            else:
                mensagem_erro = "Complete o CAPTCHA para continuar"
            self.msg_label.configure(text=mensagem_erro, text_color="#e74c3c")
            self.captcha_error = True
            return
            
        if campos_vazios:
            mensagem_erro = "Preencha os campos de " + ", ".join(campos_vazios)
            self.msg_label.configure(text=mensagem_erro, text_color="#e74c3c")
            return

        users = carregar_dados_banco(ARQUIVO_USUARIOS)
        user = next((u for u in users if u["usuario"] == usuario), None)
        
        if not user and not any(u["senha"] == senha for u in users):
            self.msg_label.configure(text="Usuário e senha incorretos", text_color="#e74c3c")
            messagebox.showerror("Erro", "Usuário e senha estão incorretos ou não encontrados. Verifique as informações e tente novamente.")
            self.usuario_entry.configure(border_color="#ffcccc")
            self.senha_entry.configure(border_color="#ffcccc")
            self.usuario_asterisk.configure(text="×")
            self.senha_asterisk.configure(text="×")
            return
            
        if not user:
            self.msg_label.configure(text="Usuário incorreto", text_color="#e74c3c")
            messagebox.showerror("Erro", "Usuário incorreto. Verifique as informações e tente novamente.")
            self.usuario_entry.configure(border_color="#ffcccc")
            self.usuario_asterisk.configure(text="×")
            return
            
        if user["senha"] != senha:
            self.msg_label.configure(text="Senha incorreta", text_color="#e74c3c")
            messagebox.showerror("Erro", "Senha incorreta. Verifique as informações e tente novamente.")
            self.senha_entry.configure(border_color="#ffcccc")
            self.senha_asterisk.configure(text="×")
            return

        self.msg_label.configure(text="")
        
        if user["filtro"] in ["Administrativo", "Desenvolvimento"]:
            self.perguntar_sistema(user)
        else:
            self.abrir_principal(user)

    def _perguntar_sistema(self, user):
        win = ctk.CTkToplevel(self)
        win.title("Selecionar Sistema")
        win.geometry("400x200")
        win.resizable(False, False)
        win.grab_set()
        win.transient(self)
        win.focus_set()
        win.after(200, lambda: win.iconbitmap("icones//logo.ico"))

        ctk.CTkLabel(
            win, 
            text="Selecione o sistema que deseja acessar:", 
            font=("Roboto", 16)
        ).pack(pady=20)

        btn_frame = ctk.CTkFrame(win, fg_color="transparent")
        btn_frame.pack(pady=20)

        btn_cadastro = ctk.CTkButton(
            btn_frame, 
            text="Cadastros", 
            command=lambda: [win.destroy(), self.abrir_principal(user)],
            fg_color="#047194", 
            hover_color="#008ab6", 
            text_color="#fff",
            font=("Roboto", 16, "bold"), 
            corner_radius=12, 
            height=40, 
            width=120
        )
        btn_cadastro.pack(side="left", padx=20)

        btn_locacao = ctk.CTkButton(
            btn_frame, 
            text="Locação", 
            command=lambda: messagebox.showinfo("Aviso", "Sistema de locação ainda não está disponível."),
            fg_color="#9b59b6", 
            hover_color="#8e44ad", 
            text_color="#fff",
            font=("Roboto", 16, "bold"), 
            corner_radius=12, 
            height=40,  
            width=120,
            state="disabled"
        )
        btn_locacao.pack(side="left", padx=20)

    def _abrir_registro(self):
        win = TelaRegistroUsuario(self)
        win.grab_set()

    def _abrir_principal(self, user):
        self.withdraw()
        win = TelaPrincipal(self, user)
        win.grab_set()

    def build_login(self):
        for widget in self.winfo_children():
            widget.destroy()
        
        main_container = ctk.CTkFrame(self, fg_color="#F4F1F1")
        main_container.pack(fill="both", expand=True)    

        img_frame = ctk.CTkFrame(main_container, fg_color="#202020")
        img_frame.place(relx=0, rely=0, relwidth=0.55, relheight=1)

        try:
            img = Image.open("layout_login.png")
            def resize_img(event):
                w = int(img_frame.winfo_width())
                h = int(img_frame.winfo_height())
                if w > 0 and h > 0:
                    img_resized = img.resize((w, h))
                    self.bg_img = ctk.CTkImage(light_image=img_resized, size=(w, h))
                    if hasattr(self, "img_label"):
                        self.img_label.configure(image=self.bg_img)
                    else:
                        self.img_label = ctk.CTkLabel(img_frame, image=self.bg_img, text="")
                        self.img_label.place(relx=0, rely=0, relwidth=1, relheight=1)
            img_frame.bind("<Configure>", resize_img)
        except Exception as e:
            pass
        
        login_container = ctk.CTkFrame(main_container, fg_color="transparent")
        login_container.place(relx=0.55, rely=0, relwidth=0.45, relheight=1)

        login_frame = ctk.CTkFrame(
            login_container,
            fg_color="#ffffff",
            corner_radius=18,
            border_color="#D5D5D5",
            border_width=2,
            width=400, height=600
        )
        login_frame.place(relx=0.5, rely=0.5, anchor="center")

        ctk.CTkLabel(
            login_frame, 
            text="Bem-vindo!", 
            font=("Roboto", 55, "bold"), 
            text_color="#464646"
        ).pack(pady=(20, 10))

        ctk.CTkFrame(login_frame, fg_color="transparent", height=100).pack()

        usuario_label = ctk.CTkLabel(
            login_frame, 
            text="Usuário*", 
            font=("Roboto", 12), 
            anchor="w", 
            justify="left",

        )
        usuario_label.pack(pady=(0,0), padx=(130), anchor="w") 
        self.usuario_entry = ctk.CTkEntry(
            login_frame, 
            textvariable=self.usuario_var, 
            width=280, 
            height=38, 
            corner_radius=10, 
            font=("Roboto", 16)
        )
        self.usuario_entry.pack(padx=130)
        self.usuario_asterisk = ctk.CTkLabel(
            login_frame, 
            text="", 
            text_color="red", 
            font=("Roboto", 24, "bold"), 
            fg_color="#f9f9fa"
        )
        self.usuario_asterisk.place(relx=0.72, rely=0.39)

        senha_label = ctk.CTkLabel(
            login_frame, 
            text="Senha*", 
            font=("Roboto", 12), 
            anchor="w", 
            justify="left",
        )
        senha_label.pack(pady=(8,0), padx=(130), anchor="w") 
        self.senha_entry = ctk.CTkEntry(
            login_frame, 
            textvariable=self.senha_var, 
            show="●", 
            width=280, 
            height=38, 
            corner_radius=10, 
            font=("Roboto", 13)
        )
        self.senha_entry.pack(padx=28)
        self.senha_asterisk = ctk.CTkLabel(
            login_frame, 
            text="", 
            text_color="red", 
            font=("Roboto", 24, "bold"), 
            fg_color="#f9f9fa"
        )
        self.senha_asterisk.place(relx=0.72, rely=0.52)

        captcha_main_frame = ctk.CTkFrame(
            login_frame,
            fg_color="#f0f8ff",
            border_width=2,
            border_color="#A3A3A3",
            corner_radius=8
        )
        captcha_main_frame.pack(pady=(15, 5), fill="x", padx=130)

        captcha_inner_frame = ctk.CTkFrame(captcha_main_frame, fg_color="transparent")
        captcha_inner_frame.pack(padx=10, pady=8)

        self.captcha_checkbox = ctk.CTkCheckBox(
            captcha_inner_frame,
            text="Não sou um robô",  
            variable=self.captcha_var,
            width=58,  
            height=58,  
            corner_radius=5,  
            border_width=2,
            command=self.verificar_captcha,
            font=("Roboto", 18)  
        )
        self.captcha_checkbox.pack(side="left", padx=(0, 10))

        icon_frame = ctk.CTkFrame(captcha_inner_frame, fg_color="transparent")
        icon_frame.pack(side="left", padx=(10, 0), pady=8)

        try:
            captcha_img = Image.open("icones//captcha.png").resize((30, 30), Image.LANCZOS)
            self.captcha_icon = ctk.CTkImage(light_image=captcha_img, dark_image=captcha_img, size=(30, 30))
            ctk.CTkLabel(icon_frame, image=self.captcha_icon, text="").pack()
        except:
            ctk.CTkLabel(icon_frame).pack()

        ctk.CTkLabel(
            icon_frame,
            text="TecTCHA",
            font=("Roboto", 9),
            text_color="black"
        ).pack()

        self.msg_label = ctk.CTkLabel(
            login_frame, 
            text="", 
            text_color="#e74c3c", 
            font=("Roboto", 13, "bold")
        )
        self.msg_label.pack(pady=(10,0))

        btn_frame = ctk.CTkFrame(login_frame, fg_color="transparent")
        btn_frame.pack(pady=(0,0))
        ctk.CTkButton(
            btn_frame,
            text="Entrar", 
            command=self.login,
            fg_color="#047194", 
            hover_color="#008ab6", 
            text_color="#fff",
            font=("Roboto", 18, "bold"), 
            corner_radius=12, 
            height=44, 
            width=120,
            border_width=2,
            border_color="#045975"
        ).pack(side="left", padx=(0, 10))
        ctk.CTkButton(
            btn_frame, 
            text="Registrar-se", 
            command=self.abrir_registro,
            fg_color="#047194", 
            hover_color="#008ab6", 
            text_color="#fff",
            font=("Roboto", 18, "bold"), 
            corner_radius=12, 
            height=44, 
            width=120,
            border_width=2,
            border_color="#045975"
        ).pack(side="left")
        

        esqueceu_frame = ctk.CTkFrame(login_frame, fg_color="transparent")
        esqueceu_frame.pack(pady=(15, 10))
        ctk.CTkLabel(
            esqueceu_frame,
            text="Esqueceu a senha?",
            font=("Roboto", 12),
            text_color="#222"
        ).pack(side="left")
        ctk.CTkButton(
            esqueceu_frame,
            text="Clique aqui",
            fg_color="transparent",
            hover_color="#e0f7fa",
            text_color="#001cd4",
            font=("Roboto", 12, "underline"),
            corner_radius=8,
            width=1,
            height=1,
            command=lambda: None
        ).pack(side="left", padx=(4,0))

class TelaRegistroUsuario(ctk.CTkToplevel):
    def __init__(self, master, dados_usuario=None):
        super().__init__(master)
        self.title("Registrar-se" if not dados_usuario else "Alterar Dados do Usuário")
        self.geometry("800x670")
        self.resizable(False, False)
        self.configure(fg_color="#f7f7fb")
        self.foto_path = SEM_IMAGEM
        self.dados_usuario = dados_usuario
        self.after(200, lambda: self.iconbitmap("icones//logo.ico"))
        
        self.vars = {
            "nome": ctk.StringVar(value=tratar_valor_vazio(dados_usuario.get("nome", "")) if dados_usuario else ""),
            "data_nasc": ctk.StringVar(value=tratar_valor_vazio(dados_usuario.get("data_nasc", "")) if dados_usuario else ""),
            "departamento": ctk.StringVar(value=tratar_valor_vazio(dados_usuario.get("departamento", "")) if dados_usuario else ""),
            "setor": ctk.StringVar(value=tratar_valor_vazio(dados_usuario.get("setor", "")) if dados_usuario else ""),
            "usuario": ctk.StringVar(value=tratar_valor_vazio(dados_usuario.get("usuario", "")) if dados_usuario else ""),
            "senha": ctk.StringVar(value=""),
            "email": ctk.StringVar(value=tratar_valor_vazio(dados_usuario.get("email", "")) if dados_usuario else ""),
            "filtro": ctk.StringVar(value=tratar_valor_vazio(dados_usuario.get("filtro", "Padrão")) if dados_usuario else "Padrão")
        }

        ctk.CTkLabel(
            self, 
            text="Cadastro de Usuario" if not dados_usuario else "Alterar Dados do Usuario", 
            font=("Roboto", 22, "bold"), 
            text_color="#3a3a3a"
        ).pack(pady=(18, 10))

        frame_principal = ctk.CTkFrame(self, fg_color="#ffffff", corner_radius=18)
        frame_principal.pack(padx=24, pady=10, fill="both", expand=True)

        frame_principal.grid_columnconfigure(0, weight=1)
        frame_principal.grid_columnconfigure(1, weight=0)
        frame_principal.grid_columnconfigure(2, weight=1)
        frame_principal.grid_rowconfigure(0, weight=1)

        form_frame = ctk.CTkFrame(frame_principal, fg_color="transparent")
        form_frame.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")

        campos = [
            ("Nome completo*", "nome"),
            ("Email*", "email"),
            ("Data de Nascimento (dd/mm/aaaa)*", "data_nasc"),
            ("Departamento*", "departamento"),
            ("Setor*", "setor"),
            ("Filtro*", "filtro")
        ]
        
        if not dados_usuario:
            campos.append(("Usuário*", "usuario"))
            campos.append(("Senha*", "senha"))

        for label, campo in campos:
            ctk.CTkLabel(
                form_frame, 
                text=label, 
                font=("Roboto", 13)
            ).pack(pady=(8,0), anchor="w", padx=48)
            
            if campo == "filtro":
                opcoes = ["Padrão", "Analista", "Administrativo", "Logística", "Manutenção", "Desenvolvimento"]
                entry = ctk.CTkOptionMenu(
                    form_frame, 
                    variable=self.vars[campo], 
                    values=opcoes,
                    width=300, 
                    height=30, 
                    corner_radius=10, 
                    font=("Roboto", 12)
                )
                entry.pack(pady=(0,2), padx=10)
            else:
                entry = ctk.CTkEntry(
                    form_frame, 
                    textvariable=self.vars[campo], 
                    show="*" if campo=="senha" else None, 
                    width=300, 
                    height=30, 
                    corner_radius=10, 
                    font=("Roboto", 12)
                )
                entry.pack(pady=(0,2), padx=10)

        photo_preview_frame = ctk.CTkFrame(frame_principal, fg_color="transparent")
        photo_preview_frame.grid(row=0, column=1, padx=20, pady=20, sticky="n")

        ctk.CTkLabel(
            photo_preview_frame, 
            text="Pré-visualização da Foto", 
            font=("Roboto", 13)
        ).pack(pady=(0,5))

        self.preview_moldura = ctk.CTkFrame(
            photo_preview_frame, 
            width=220, 
            height=220,
            corner_radius=20, 
            fg_color="#F0F8FF", 
            border_width=4, 
            border_color="#4F8CFF"
        )
        self.preview_moldura.pack(pady=(0, 10))
        self.preview_moldura.pack_propagate(False)

        self.preview_label = ctk.CTkLabel(self.preview_moldura, text="", fg_color="transparent")
        self.preview_label.pack(expand=True)
        
        if dados_usuario and dados_usuario.get("foto"):
            self.carregar_preview_image(dados_usuario["foto"])
            self.foto_path = dados_usuario["foto"]
        else:
            self.carregar_preview_image(SEM_IMAGEM)

        self.foto_btn = ctk.CTkButton(
            photo_preview_frame, 
            text="Selecionar Foto", 
            command=self.selecionar_foto, 
            fg_color="#4f8cff", 
            hover_color="#357ae8", 
            text_color="#fff", 
            corner_radius=10,
            width=200, 
            height=38,
            border_width=1,
            border_color="#D5D5D5"
        )
        self.foto_btn.pack(pady=(5,0))

        btn_text = "Registrar" if not dados_usuario else "Salvar Alterações"
        ctk.CTkButton(
            frame_principal, 
            text=btn_text, 
            command=self.registrar, 
            fg_color="#4f8cff", 
            hover_color="#357ae8", 
            text_color="#fff", 
            font=("Roboto", 14, "bold"), 
            corner_radius=12, 
            height=38, 
            width=200,
            border_width=1,
            border_color="#D5D5D5"
        ).grid(row=1, column=0, columnspan=2, pady=18)

    def carregar_preview_image(self, path):
        try:
            img = Image.open(path).resize((200, 200), Image.LANCZOS)
            self.preview_img_tk = ctk.CTkImage(light_image=img, dark_image=img, size=(200, 200))
            self.preview_label.configure(image=self.preview_img_tk)
        except Exception:
            if path != SEM_IMAGEM:
                self.carregar_preview_image(SEM_IMAGEM)
            else:
                self.preview_label.configure(text="Erro ao carregar imagem", text_color="red")

    def selecionar_foto(self):
        path = filedialog.askopenfilename(filetypes=[("Imagens", "*.png;*.jpg;*.jpeg")])
        if path:
            self.foto_path = path
            self.carregar_preview_image(path)

    def registrar(self):
        dados = {k: tratar_valor_vazio(v.get()) for k, v in self.vars.items()}
        
        required_fields = ["nome", "email", "data_nasc", "departamento", "setor", "filtro"]
        if not self.dados_usuario:
            required_fields += ["usuario", "senha"]
            
        for campo in required_fields:
            if not dados.get(campo) or dados.get(campo) == "N/A":
                messagebox.showerror("Erro", f"O campo '{campo.replace('_', ' ').capitalize()}' é obrigatório.")
                return
        
        if not validar_email(dados["email"]):
            messagebox.showerror("Erro", "Email inválido.")
            return
        
        idade = calcular_idade(dados["data_nasc"])
        if idade is None:
            messagebox.showerror("Erro", "Data de nascimento inválida. Use o formato dd/mm/aaaa.")
            return
        dados["idade"] = idade

        if not self.dados_usuario:
            if not self.foto_path:
                self.foto_path = SEM_IMAGEM
                os.makedirs(os.path.dirname(SEM_IMAGEM), exist_ok=True)
                if not os.path.exists(SEM_IMAGEM):
                    placeholder_img = Image.new('RGB', (50, 50), color = 'black')
                    placeholder_img.save(SEM_IMAGEM)

            users = carregar_dados_banco(ARQUIVO_USUARIOS)
            if any(u["usuario"] == dados["usuario"] for u in users):
                messagebox.showerror("Erro", "Usuário já existe.")
                return
            
            foto_dir = "fotos"
            os.makedirs(foto_dir, exist_ok=True)
            foto_dest = os.path.join(foto_dir, f'{dados["usuario"]}.png')
            
            Image.open(self.foto_path).resize((50,50), Image.LANCZOS).save(foto_dest)
            
            dados["foto"] = foto_dest
            users.append(dados)
            salvar_dados_banco(ARQUIVO_USUARIOS, users)
            messagebox.showinfo("Sucesso", "Usuário registrado!")
            self.destroy()
        
        else:
            users = carregar_dados_banco(ARQUIVO_USUARIOS)
            updated = False
            for i, user in enumerate(users):
                if user["usuario"] == self.dados_usuario["usuario"]:
                    for key in dados:
                        if key != "senha" or dados["senha"]:
                            users[i][key] = dados[key]
                    
                    if self.foto_path != self.dados_usuario["foto"]:
                        if os.path.exists(self.dados_usuario["foto"]) and self.dados_usuario["foto"] != SEM_IMAGEM:
                            try:
                                os.remove(self.dados_usuario["foto"])
                            except OSError as e:
                                print(f"Erro ao remover foto antiga: {e}")
                        
                        foto_dir = "fotos"
                        os.makedirs(foto_dir, exist_ok=True)
                        foto_dest = os.path.join(foto_dir, f'{self.dados_usuario["usuario"]}.png')
                        Image.open(self.foto_path).resize((50,50), Image.LANCZOS).save(foto_dest)
                        users[i]["foto"] = foto_dest
                    
                    updated = True
                    break
            
            if updated:
                salvar_dados_banco(ARQUIVO_USUARIOS, users)
                messagebox.showinfo("Sucesso", "Informações do usuário salvas com sucesso!")
                self.master.atualizar_info_display_usuario()
                self.destroy()
            else:
                messagebox.showerror("Erro", "Usuário não encontrado para atualização.")

class TelaCadastroPessoa(ctk.CTkToplevel):
    def __init__(self, master, pessoa=None):
        super().__init__(master)
        self.title("Cadastrar Pessoa" if not pessoa else "Editar Pessoa")
        self.geometry("700x750")
        self.resizable(False, False)
        self.configure(fg_color="#f7f7fb")
        self.pessoa_original = pessoa
        self.tipo_var = ctk.StringVar(value="cliente")
        self.campos_frame = None
        self.vars = {}
        self.after(200, lambda: self.iconbitmap("icones//logo.ico"))
        
        self.cep_disponivel = True
        try:
            import requests
            from COD_CEP import CEPService
        except ImportError:
            self.cep_disponivel = False
            messagebox.showwarning(
                "Aviso", 
                "Funcionalidade de CEP não disponível (módulo requests não instalado). "
                "Instale com: pip install requests"
            )

        ctk.CTkLabel(
            self, 
            text="Cadastro de Pessoas" if not pessoa else "Editar Pessoa", 
            font=("Roboto", 22, "bold"), 
            text_color="#3a3a3a"
        ).pack(pady=(18, 10))

        frame_principal = ctk.CTkScrollableFrame(self, fg_color="#ffffff", corner_radius=18)
        frame_principal.pack(padx=24, pady=10, fill="both", expand=True)
        
        if not pessoa:
            ctk.CTkLabel(
                frame_principal, 
                text="Tipo*", 
                font=("Roboto", 13)
            ).pack(pady=(8,0), anchor="w", padx=158)
            tipo_menu = ctk.CTkOptionMenu(
                frame_principal, 
                variable=self.tipo_var, 
                values=["cliente", "funcionario", "fornecedor"],
                command=self.atualizar_campos,
                width=300, 
            height=32, 
                corner_radius=10, 
                font=("Roboto", 12)
            )
            tipo_menu.pack(pady=(0,2), padx=18)
        else:
            ctk.CTkLabel(frame_principal, text=f"Tipo: {pessoa['tipo']}", font=("Roboto", 14, "bold")).pack(pady=10)
            self.tipo_var.set(pessoa["tipo"])
        
        self.campos_container = ctk.CTkFrame(frame_principal, fg_color="transparent")
        self.campos_container.pack(fill="both", expand=True, pady=10)
        
        self.atualizar_campos()
        
        btn_text = "Cadastrar" if not pessoa else "Salvar Alterações"
        ctk.CTkButton(
            frame_principal, 
            text=btn_text, 
            command=self.cadastrar, 
            fg_color="#047194", 
            hover_color="#008ab6", 
            text_color="#fff", 
            font=("Roboto", 18, "bold"), 
            corner_radius=12, 
            height=38, 
            width=110,
            border_color="#045975",
            border_width=2
        ).pack(pady=18)

    def buscar_cep(self):
        cep = self.vars["cep"].get()
        if not validar_cep(cep):
            messagebox.showerror("Erro", "CEP inválido. Formato esperado: 00000-000 ou 00000000")
            return
            
        try:
            cep_limpo = re.sub(r'[^0-9]', '', cep)
            from COD_CEP import CEPService
            resultado = CEPService.buscar_cep(cep_limpo)
            
            if resultado.get('erro'):
                if resultado['tipo'] == 'nao_encontrado':
                    msg = "CEP não encontrado. Por favor, preencha o endereço manualmente."
                else:
                    msg = f"Não foi possível buscar o CEP: {resultado['mensagem']}"
                
                messagebox.showwarning("CEP", msg)
                self.vars["endereco"].set("N/A")
                self.vars["bairro"].set("N/A")
                self.vars["cidade"].set("N/A")
                self.vars["estado"].set("N/A")
                return
                
            self.vars["endereco"].set(tratar_valor_vazio(resultado.get("logradouro", "")))
            self.vars["bairro"].set(tratar_valor_vazio(resultado.get("bairro", "")))
            self.vars["cidade"].set(tratar_valor_vazio(resultado.get("cidade", "")))
            self.vars["estado"].set(tratar_valor_vazio(resultado.get("estado", "")))
            
            self.focus_set()
            for widget in self.winfo_children():
                if isinstance(widget, ctk.CTkEntry) and widget._textvariable == self.vars["numero"]:
                    widget.focus()
                    break
                    
        except ImportError:
            messagebox.showerror("Erro", "Serviço de CEP não disponível. Contate o suporte.")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha inesperada ao buscar CEP: {str(e)}")

    def atualizar_campos(self, *args):
        if self.campos_frame:
            self.campos_frame.destroy()
            
        self.campos_frame = ctk.CTkFrame(self.campos_container, fg_color="transparent")
        self.campos_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        tipo = self.tipo_var.get()
        self.vars = {}
        
        campos = [("Nome Completo*", "nome")]
        
        if tipo == "funcionario":
            campos += [
                ("CPF*", "cpf"),
                ("Data de Nascimento (dd/mm/aaaa)*", "data_nasc"),
                ("Departamento*", "departamento"),
                ("Setor*", "setor"),
                ("Data de Admissão (dd/mm/aaaa)*", "data_admissao")
            ]
        elif tipo == "fornecedor":
            campos += [
                ("CNPJ*", "cnpj"),
                ("Tipo de Fornecimento*", "tipo_fornecimento"),
                ("CEP*", "cep"),
                ("Endereço", "endereco"),
                ("Número", "numero"),
                ("Bairro", "bairro"),
                ("Cidade", "cidade"),
                ("Estado", "estado")
            ]
        else: 
            campos += [
                ("CPF*", "cpf"),
                ("Data de Nascimento (dd/mm/aaaa)*", "data_nasc"),
                ("CEP*", "cep"),
                ("Endereço", "endereco"),
                ("Número", "numero"),
                ("Bairro", "bairro"),
                ("Cidade", "cidade"),
                ("Estado", "estado")
            ]
        
        dados_existentes = self.pessoa_original if self.pessoa_original else {}
        
        for label, campo in campos:
            ctk.CTkLabel(self.campos_frame, text=label, font=("Roboto", 13)).pack(pady=(8,0), anchor="w", padx=150)

            if campo == "cep":
                cep_frame = ctk.CTkFrame(self.campos_frame, fg_color="transparent")
                cep_frame.pack(fill="x", padx=150)
                
                self.vars[campo] = ctk.StringVar(value=tratar_valor_vazio(dados_existentes.get(campo, "")))
                cep_entry = ctk.CTkEntry(
                    cep_frame, 
                    textvariable=self.vars[campo], 
                    width=240, 
                    height=32, 
                    corner_radius=10, 
                    font=("Roboto", 12),
                    placeholder_text="00000-000"
                )
                cep_entry.pack(side="left", padx=(0,5))
                
                self.buscar_cep_btn = ctk.CTkButton(
                    cep_frame, 
                    text="🔍", 
                    font=("Roboto", 20),
                    width=40, 
                    height=32, 
                    corner_radius=10,
                    command=self.buscar_cep,
                    fg_color="#047194", 
                    hover_color="#008ab6",
                    border_color="#045975",
                    border_width=2
                )
                self.buscar_cep_btn.pack(side="left")
                
                if not self.cep_disponivel:
                    self.buscar_cep_btn.configure(state="disabled", fg_color="#cccccc")
            else:
                self.vars[campo] = ctk.StringVar(value=tratar_valor_vazio(dados_existentes.get(campo, "")))
                entry = ctk.CTkEntry(
                    self.campos_frame, 
                    textvariable=self.vars[campo], 
                    width=300, 
                    height=32, 
                    corner_radius=10, 
                    font=("Roboto", 12)
                )
                entry.pack(pady=(0,2), padx=18)

    def cadastrar(self):
        tipo = self.tipo_var.get()
        dados = {k: tratar_valor_vazio(v.get()) for k, v in self.vars.items()}
        dados["tipo"] = tipo
        
        erros = []
        if tipo == "funcionario":
            if not dados.get("cpf") or dados.get("cpf") == "N/A": erros.append("CPF")
            if not dados.get("departamento") or dados.get("departamento") == "N/A": erros.append("Departamento")
            if not dados.get("setor") or dados.get("setor") == "N/A": erros.append("Setor")
        elif tipo == "fornecedor":
            if not dados.get("cnpj") or dados.get("cnpj") == "N/A": erros.append("CNPJ")
            if not dados.get("tipo_fornecimento") or dados.get("tipo_fornecimento") == "N/A": erros.append("Tipo de Fornecimento")
            if not dados.get("cep") or dados.get("cep") == "N/A": erros.append("CEP")
        else: 
            if not dados.get("cpf") or dados.get("cpf") == "N/A": erros.append("CPF")
            if not dados.get("cep") or dados.get("cep") == "N/A": erros.append("CEP")
        
        if not dados.get("nome") or dados.get("nome") == "N/A":
            erros.append("Nome")
        
        if erros:
            campos = "\n".join([f"• {campo}" for campo in erros])
            messagebox.showerror("Erro", f"Por favor, preencha os campos obrigatórios:\n{campos}")
            return
        
        if "cpf" in dados and dados["cpf"] != "N/A" and not validar_cpf(dados["cpf"]):
            messagebox.showerror("Erro", "CPF inválido.")
            return
        if "cnpj" in dados and dados["cnpj"] != "N/A" and not validar_cnpj(dados["cnpj"]):
            messagebox.showerror("Erro", "CNPJ inválido.")
            return
        if "cep" in dados and dados["cep"] != "N/A" and not validar_cep(dados["cep"]):
            messagebox.showerror("Erro", "CEP inválido.")
            return
        
        if "data_nasc" in dados and dados["data_nasc"] != "N/A":
            idade = calcular_idade(dados["data_nasc"])
            if idade is None:
                messagebox.showerror("Erro", "Data de nascimento inválida. Use o formato dd/mm/aaaa.")
                return
            dados["idade"] = idade

        arquivo = ARQUIVO_PESSOAS[tipo]
        pessoas = carregar_dados_banco(arquivo)
        
        if self.pessoa_original:
            documento_original = self.pessoa_original.get("cpf") or self.pessoa_original.get("cnpj")
            documento_novo = dados.get("cpf") or dados.get("cnpj")
            
            if documento_novo != documento_original:
                pessoa_existente, _ = buscar_pessoa_por_documento(documento_novo)
                if pessoa_existente:
                    messagebox.showerror("Erro", "Documento já cadastrado para outra pessoa.")
                    return
        else:
            documento = dados.get("cpf") or dados.get("cnpj")
            pessoa_existente, _ = buscar_pessoa_por_documento(documento)
            if pessoa_existente:
                messagebox.showerror("Erro", "Documento já cadastrado.")
                return
        
        if self.pessoa_original:
            documento_original = self.pessoa_original.get("cpf") or self.pessoa_original.get("cnpj")
            pessoas = [p for p in pessoas if (p.get("cpf") != documento_original and p.get("cnpj") != documento_original)]
        
        pessoas.append(dados)
        salvar_dados_banco(arquivo, pessoas)
        messagebox.showinfo("Sucesso", "Dados salvos com sucesso!" if self.pessoa_original else "Pessoa cadastrada com sucesso!")
        self.master.master.atualizar_lista()
        self.destroy()

class TelaExportar(ctk.CTkToplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Exportar Dados")
        self.geometry("300x200")
        self.resizable(False, False)
        self.configure(fg_color="#f7f7fb")
        self.tipo_var = ctk.StringVar(value="cliente")
        self.after(200, lambda: self.iconbitmap("icones//logo.ico"))

        ctk.CTkLabel(self, text="Exportar Dados", font=("Roboto", 18, "bold")).pack(pady=20)

        ctk.CTkLabel(self, text="Selecione o tipo de pessoa:").pack()
        tipo_menu = ctk.CTkOptionMenu(
            self, 
            variable=self.tipo_var, 
            values=["cliente", "funcionario", "fornecedor"],
            width=200,
            height=30,
            corner_radius=8
        )
        tipo_menu.pack(pady=5)
        
        ctk.CTkButton(
            self, 
            text="Exportar para Excel", 
            command=self.exportar_excel,
            fg_color="#2ecc71",
            hover_color="#27ae60",
            text_color="#fff",
            font=("Sego UI", 14, "bold"),
            corner_radius=12,
            height=40,
            width=200
        ).pack(pady=20)

    def exportar_excel(self):
        tipo = self.tipo_var.get()
        arquivo = ARQUIVO_PESSOAS[tipo]
        pessoas = carregar_dados_banco(arquivo)
        
        if not pessoas:
            messagebox.showinfo("Exportar", "Nenhum dado para exportar.")
            return
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Salvar arquivo como"
        )
        
        if not file_path:  
            return
        
        df = pd.DataFrame(pessoas)
        
        df = df.fillna("N/A")
        
        if 'tipo' in df.columns:
            df = df.drop(columns=['tipo'])
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = tipo.capitalize()
            
            headers = list(df.columns)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for row_num, row in enumerate(df.values, 2):
                for col_num, value in enumerate(row, 1):
                    valor_tratado = tratar_valor_vazio(value)
                    ws.cell(row=row_num, column=col_num, value=valor_tratado)
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=1, max_row=len(pessoas)+1, max_col=len(headers)):
                for cell in row:
                    cell.border = thin_border
            
            wb.save(file_path)
            messagebox.showinfo("Sucesso", f"Dados exportados para {file_path}")
            self.destroy()
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao exportar: {str(e)}")

class TelaImportar(ctk.CTkToplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Importar Dados")
        self.geometry("500x300")
        self.resizable(False, False)
        self.configure(fg_color="#f7f7fb")
        self.file_path = ""
        self.after(200, lambda: self.iconbitmap("icones//logo.ico"))

        ctk.CTkLabel(self, text="Importar Dados", font=("Roboto", 18, "bold")).pack(pady=20)

        file_frame = ctk.CTkFrame(self, fg_color="transparent")
        file_frame.pack(fill="x", padx=50, pady=10)

        self.file_var = ctk.StringVar()
        ctk.CTkLabel(file_frame, text="Arquivo:").pack(side="left", padx=(0, 10))
        self.file_entry = ctk.CTkEntry(
            file_frame, 
            textvariable=self.file_var, 
            width=300, 
            height=30, 
            corner_radius=10
        )
        self.file_entry.pack(side="left", padx=(0, 5))

        browse_btn = ctk.CTkButton(
            file_frame, 
            text="💾", 
            font=("Roboto", 23, "bold"),
            width=40, 
            height=30, 
            corner_radius=10,
            command=self.browse_file,
            fg_color="#4f8cff", 
            hover_color="#357ae8"
        )
        browse_btn.pack(side="left")

        ctk.CTkButton(
            self, 
            text="⬇️Baixar Modelo", 
            font=("Roboto", 16, "bold"),
            command=self.baixar_modelo,
            fg_color="#2ecc71",
            hover_color="#27ae60",
            text_color="#fff",
            corner_radius=10,
            height=30
        ).pack(pady=10)

        ctk.CTkButton(
            self, 
            text="Importar➡️", 
            command=self.importar,
            fg_color="#2ecc71",
            hover_color="#27ae60",
            text_color="#fff",
            font=("Roboto", 18, "bold"),
            corner_radius=12,
            height=40,
            width=200
        ).pack(pady=20)

    def browse_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.file_var.set(file_path)
            self.file_path = file_path

    def baixar_modelo(self):
        modelo_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv")],
            title="Salvar modelo como"
        )
        if modelo_path:
            campos = {
                "funcionario": ["nome", "cpf", "data_nasc", "departamento", "setor", "data_admissao", "email"],
                "fornecedor": ["nome", "cnpj", "tipo_fornecimento", "cep", "endereco", "numero", "bairro", "cidade", "estado", "email"],
                "cliente": ["nome", "cpf", "data_nasc", "cep", "endereco", "numero", "bairro", "cidade", "estado", "email"]
            }
            
            with open(modelo_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                header = ["tipo"] + list(set([campo for sublist in campos.values() for campo in sublist]))
                writer.writerow(header)
                
                writer.writerow(["funcionario", "João Silva", "123.456.789-00", "01/01/1990", "TI", "Desenvolvimento", "10/02/2020", "joao@empresa.com"])
                writer.writerow(["fornecedor", "Fornecedor ABC", "12.345.678/0001-90", "Materiais", "12345-678", "Rua Exemplo", "123", "Centro", "Cidade", "SP", "fornecedor@abc.com"])
                writer.writerow(["cliente", "Maria Souza", "987.654.321-00", "15/05/1985", "54321-000", "Av. Principal", "456", "Jardim", "Cidade", "RJ", "maria@email.com"])
                
            messagebox.showinfo("Modelo", f"Modelo salvo em: {modelo_path}")

    def importar(self):
        if not self.file_path:
            messagebox.showerror("Erro", "Selecione um arquivo CSV или Excel.")
            return

        try:
            if self.file_path.endswith('.csv'):
                df = pd.read_csv(self.file_path, encoding='utf-8')
            elif self.file_path.endswith('.xlsx'):
                df = pd.read_excel(self.file_path)
            else:
                messagebox.showerror("Erro", "Formato não suportado. Use CSV или Excel.")
                return
            
            campos_por_tipo = {
                "funcionario": {
                    "obrigatorios": ["nome", "cpf", "data_nasc", "departamento", "setor", "data_admissao"],
                    "opcionais": ["email"]
                },
                "fornecedor": {
                    "obrigatorios": ["nome", "cnpj", "tipo_fornecimento", "cep", "endereco", "numero", "bairro", "cidade", "estado"],
                    "opcionais": ["email"]
                },
                "cliente": {
                    "obrigatorios": ["nome", "cpf", "data_nasc", "cep", "endereco", "numero", "bairro", "cidade", "estado"],
                    "opcionais": ["email"]
                }
            }
            
            erros = []
            dados_por_tipo = {tipo: [] for tipo in campos_por_tipo}
            
            for idx, registro in df.iterrows():
                registro_dict = registro.to_dict()
                tipo = str(registro_dict.get('tipo', '')).lower().strip()
                
                if tipo not in campos_por_tipo:
                    erros.append(f"Linha {idx+1}: Tipo '{tipo}' inválido")
                    continue
                
                campos_obrigatorios = campos_por_tipo[tipo]["obrigatorios"]
                campos_opcionais = campos_por_tipo[tipo]["opcionais"]
                
                campos_faltantes = [campo for campo in campos_obrigatorios if not str(registro_dict.get(campo, '')).strip()]
                if campos_faltantes:
                    erros.append(f"Linha {idx+1}: Campos obrigatórios faltando: {', '.join(campos_faltantes)}")
                    continue
                
                registro_filtrado = {"tipo": tipo}
                
                for campo in campos_obrigatorios + campos_opcionais:
                    valor = str(registro_dict.get(campo, '')).strip()
                    registro_filtrado[campo] = tratar_valor_vazio(valor)
                
                documento = registro_filtrado.get('cpf') or registro_filtrado.get('cnpj')
                if documento and documento != "N/A":
                    pessoa_existente, _ = buscar_pessoa_por_documento(documento)
                    if pessoa_existente:
                        erros.append(f"Linha {idx+1}: Documento {documento} já cadastrado")
                        continue
                
                if 'data_nasc' in registro_filtrado and registro_filtrado['data_nasc'] != "N/A":
                    idade = calcular_idade(registro_filtrado['data_nasc'])
                    if idade is None:
                        erros.append(f"Linha {idx+1}: Data de nascimento inválida")
                        continue
                    registro_filtrado['idade'] = idade
                else:
                    registro_filtrado['idade'] = "N/A"
                
                dados_por_tipo[tipo].append(registro_filtrado)
            
            if erros:
                message = "\n".join(erros)
                messagebox.showerror("Erros", f"Corrija os erros:\n{message}")
                return
            
            for tipo, dados in dados_por_tipo.items():
                if dados:
                    arquivo = ARQUIVO_PESSOAS[tipo]
                    dados_existentes = carregar_dados_banco(arquivo)
                    dados_existentes.extend(dados)
                    salvar_dados_banco(arquivo, dados_existentes)
            
            messagebox.showinfo("Sucesso", "Dados importados com sucesso!")
            self.master.atualizar_lista()
            self.destroy()
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao importar: {str(e)}")

class TelaPrincipal(ctk.CTkToplevel):
    def __init__(self, master, user):
        super().__init__(master)
        self.title("Sistema de Cadastros")
        self.state("zoomed")
        self.configure(fg_color="#f0f0f0")
        self.user = user
        self.after(200, lambda: self.iconbitmap("icones//logo.ico"))
        
        self.grid_rowconfigure(2, weight=1)
        self.grid_columnconfigure(0, weight=1)
        
        self.criar_header()
        
        self.criar_divisoria()
        
        self.criar_body()
        
        self.atualizar_relogio()
        
        # Construir a estrutura das abas primeiro
        self.construir_abas()
        # Depois preencher com os dados
        self.atualizar_lista()

    def criar_moldura_usuario(self, parent):
        main_frame = ctk.CTkFrame(parent, fg_color="transparent")
        main_frame.pack(pady=(10, 20), padx=10, fill="x")
        
        moldura_frame = ctk.CTkFrame(
            main_frame,
            fg_color="#047194",
            border_color="#025C7A",
            border_width=4,
            corner_radius=20,
            width=220,
            height=220
        )
        moldura_frame.pack(pady=(0, 10))
        moldura_frame.pack_propagate(False)
        
        try:
            foto_path = self.user.get("foto", SEM_IMAGEM)
            if not os.path.exists(foto_path):
                foto_path = SEM_IMAGEM
                
            user_img = Image.open(foto_path).resize((200, 200), Image.LANCZOS)
            user_photo = ctk.CTkImage(light_image=user_img, size=(200, 200))
            
            foto_label = ctk.CTkLabel(
                moldura_frame,
                image=user_photo,
                text="",
                fg_color="transparent"
            )
            foto_label.pack(expand=True, padx=10, pady=10)
        except Exception as e:
            print(f"Erro ao carregar foto: {e}")
            foto_label = ctk.CTkLabel(
                moldura_frame,
                text="📷",
                font=("Arial", 50),
                fg_color="transparent"
            )
            foto_label.pack(expand=True)

        info_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        info_frame.pack(pady=(0, 5), padx=10, fill="x")
        
        nome = self.user.get("nome", "N/A")
        nome_label = ctk.CTkLabel(
            info_frame,
            text=nome,
            font=("Roboto", 18, "bold"),
            text_color="#2C3E50",
            anchor="w",
            justify="left"
        )
        nome_label.pack(pady=(0, 5), fill="x")

        data_nasc = self.user.get("data_nasc", "N/A")
        idade = self.user.get("idade", "N/A")
        nasc_label = ctk.CTkLabel(
            info_frame,
            text=f"Nasc: {data_nasc} ({idade} anos)",
            font=("Roboto", 14),
            text_color="#566573",
            anchor="w",
            justify="left"
        )
        nasc_label.pack(pady=(0, 5), fill="x")

        departamento = self.user.get("departamento", "N/A")
        depto_label = ctk.CTkLabel(
            info_frame,
            text=f"Departamento: {departamento}",
            font=("Roboto", 14),
            text_color="#566573",
            anchor="w",
            justify="left"
        )
        depto_label.pack(pady=(0, 5), fill="x")

        setor = self.user.get("setor", "N/A")
        setor_label = ctk.CTkLabel(
            info_frame,
            text=f"Setor: {setor}",
            font=("Roboto", 14),
            text_color="#566573",
            anchor="w",
            justify="left"
        )
        setor_label.pack(pady=(0, 5), fill="x")

        return main_frame

    def criar_header(self):
        header_frame = ctk.CTkFrame(self, height=110, fg_color="#f2f2f2")
        header_frame.grid(row=0, column=0, sticky="nsew", padx=0, pady=0)
        header_frame.grid_propagate(False)
        header_frame.grid_rowconfigure(0, weight=1)
        header_frame.grid_columnconfigure(0, weight=0)
        header_frame.grid_columnconfigure(1, weight=0)
        header_frame.grid_columnconfigure(2, weight=1)
        header_frame.grid_columnconfigure(3, weight=0)
        header_frame.grid_columnconfigure(4, weight=1)
        header_frame.grid_columnconfigure(5, weight=0)

        logo_frame = ctk.CTkFrame(header_frame, width=100, height=100, fg_color="#f2f2f2")
        logo_frame.grid(row=0, column=0, padx=(15, 5), sticky="w", pady=5)
        logo_frame.grid_propagate(False)

        try:
            logo_img = Image.open("logo.png").resize((90, 90), Image.LANCZOS)
            ctk_logo_imagem = ctk.CTkImage(light_image=logo_img, size=(90, 90))
            logo_label = ctk.CTkLabel(logo_frame, image=ctk_logo_imagem, text="")
            logo_label.pack(expand=True, fill="both")
        except:
            ctk.CTkLabel(logo_frame, text="Logo", font=("Roboto", 10), text_color="black").pack(expand=True)

        user_frame = ctk.CTkFrame(header_frame, fg_color="#f2f2f2")
        user_frame.grid(row=0, column=1, sticky="w", padx=(0, 15), pady=15)

        ctk.CTkLabel(
            user_frame,
            text="Usuário: ",
            font=("Roboto", 14, "bold"),
            text_color="black"
        ).pack(side="left", padx=(0, 0))

        user_name_label = ctk.CTkLabel(
            user_frame,
            text=self.user['usuario'],
            font=("Roboto", 14, "bold"),
            text_color="red"
        )
        user_name_label.pack(side="left", padx=(0, 10))

        clock_frame = ctk.CTkFrame(
            header_frame, 
            fg_color="black", 
            border_color="#a0a0a0", 
            border_width=4, 
            corner_radius=10,
            width=220, 
            height=90
        )
        clock_frame.grid(row=0, column=3, sticky="", pady=15)
        clock_frame.grid_propagate(False)
        clock_frame.grid_rowconfigure(0, weight=3)
        clock_frame.grid_rowconfigure(1, weight=1)
        clock_frame.grid_columnconfigure(0, weight=1)

        self.time_label = ctk.CTkLabel(
            clock_frame,
            text="00:00:00",
            text_color="white",
            font=("digital-7", 40)
        )
        self.time_label.grid(row=0, column=0, sticky="nsew", padx=10, pady=(5, 0))

        self.date_label = ctk.CTkLabel(
            clock_frame,
            text="00/00/0000",
            text_color="white",
            font=("Roboto", 14, "bold")
        )
        self.date_label.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 5))

        right_frame = ctk.CTkFrame(header_frame, fg_color="#f2f2f2")
        right_frame.grid(row=0, column=4, sticky="e", padx=15, pady=15)

        buttons_frame = ctk.CTkFrame(right_frame, fg_color="#f2f2f2")
        buttons_frame.pack(side="top", anchor="e")

        btn_sobre = ctk.CTkButton(
            buttons_frame, 
            text="💡Sobre", 
            width=60,
            height=30,
            fg_color="#4f8cff", 
            hover_color="#357ae8", 
            text_color="#fff",
            font=("Segoe UI Emoji", 15, "bold"), 
            corner_radius=8, 
            border_color="#357ae8",
            border_width=2,
            command=self.abrir_sobre
        )
        btn_sobre.pack(side="left", padx=2)

        btn_ajuda = ctk.CTkButton(
            buttons_frame, 
            text="❓Ajuda", 
            width=50,
            height=30,
            fg_color="#ebb000", 
            hover_color="#f1ba13", 
            text_color="#fff",
            font=("Segoe UI Emoji", 15, "bold"), 
            corner_radius=8, 
            border_color="#d19d01",
            border_width=2,
            command=self.abrir_ajuda
        )
        btn_ajuda.pack(side="left", padx=2)

        btn_sair = ctk.CTkButton(
            buttons_frame, 
            text="❌Sair", 
            command=self.sair,
            width=48,
            height=30, 
            fg_color="#e74c3c", 
            hover_color="#c0392b", 
            text_color="#fff",
            font=("Segoe UI Emoji", 15, "bold"), 
            corner_radius=8, 
            border_width=2,
            border_color="#c0392b"
        )
        btn_sair.pack(side="left", padx=2)

    def criar_divisoria(self):
        frame_divisoria = ctk.CTkFrame(
            self, 
            height=16,
            fg_color="#b6b4b4",
            corner_radius=0
        )
        frame_divisoria.grid(row=1, column=0, sticky="ew", padx=0, pady=0)
        frame_divisoria.grid_propagate(False)

        frame_degrade = ctk.CTkFrame(
            frame_divisoria, 
            fg_color="#b6b4b4", 
            height=16,
            corner_radius=0,
        )
        frame_degrade.pack(fill="x", side="top", pady=0.003)

        ctk.CTkFrame(
            frame_degrade,
            height=8,
            fg_color="#c5c4c4",
            corner_radius=0,
        ).pack(fill="x", side="top", pady=0)

        ctk.CTkFrame(
            frame_degrade,
            height=3,
            fg_color="#b6b4b4",
            corner_radius=0,
            border_width=3,
            border_color="#b6b4b4"
        ).pack(fill="x", side="top", pady=0)

        ctk.CTkFrame(
            frame_degrade,
            height=8,
            fg_color="#A2A2A2",
            corner_radius=0
        ).pack(fill="x", side="top", pady=0.003)

    def criar_body(self):
        body_frame = ctk.CTkFrame(self, fg_color="#b7b7b7")
        body_frame.grid(row=2, column=0, sticky="nsew", padx=10, pady=(0, 10))
        body_frame.grid_rowconfigure(0, weight=1)
        body_frame.grid_columnconfigure(0, weight=1, uniform="body_cols")
        body_frame.grid_columnconfigure(1, weight=4, uniform="body_cols")

        sidebar_frame = ctk.CTkFrame(body_frame, fg_color="#f9f9f9")
        sidebar_frame.grid(row=0, column=0, sticky="nsew", padx=2, pady=2)
        sidebar_frame.grid_rowconfigure(0, weight=1)
        sidebar_frame.grid_columnconfigure(0, weight=1)

        menu_container = ctk.CTkFrame(sidebar_frame, fg_color="#f9f9f9", border_color="#c0c0c0", border_width=3, corner_radius=0)
        menu_container.pack(fill="both", expand=True, padx=0, pady=0)
        
        self.criar_moldura_usuario(menu_container)

        menu_items = [
            "⚙️ Configurações",
        ]

        for item in menu_items:
            menu_btn = ctk.CTkButton(
                menu_container,
                text=item,
                height=40,
                fg_color="#f9f9f9",
                hover_color="#e0e0e0",
                text_color="black",
                font=("Segoe UI Emoji", 16),
                anchor="w",
                command=lambda i=item: self.menu_selecionado(i)
            )
            menu_btn.pack(fill="x", pady=5, padx=10)

        self.opcao_import_export = ctk.StringVar(value="📊 Importações")
        opcoes_import_export = ["Importar", "Exportar"]
        option_menu = ctk.CTkOptionMenu(
            menu_container,
            variable=self.opcao_import_export,
            values=opcoes_import_export,
            command=self.menu_import_export_selecionado,
            height=40,
            fg_color="#f9f9f9",
            button_color="#f9f9f9",
            button_hover_color="#e0e0e0",
            text_color="black",
            font=("Segoe UI Emoji", 16),
            anchor="w",
            dropdown_fg_color="#f9f9f9",
            dropdown_hover_color="#e0e0e0",
            dropdown_text_color="black",
            dropdown_font=("Segoe UI Emoji", 16)
        )
        option_menu.pack(fill="x", pady=5, padx=10)

        agenda_btn = ctk.CTkButton(
            menu_container,
            text="📅 Agenda",
            height=40,
            fg_color="#f9f9f9",
            hover_color="#e0e0e0",
            text_color="black",
            font=("Segoe UI Emoji", 16),
            anchor="w",
            command=self.abrir_agenda
        )
        agenda_btn.pack(fill="x", pady=5, padx=10)

        content_frame = ctk.CTkFrame(body_frame, fg_color="#ffffff", border_color="#c0c0c0", border_width=3)
        content_frame.grid(row=0, column=1, sticky="nsew", padx=0, pady=0)
        content_frame.grid_rowconfigure(0, weight=1)
        content_frame.grid_columnconfigure(0, weight=1)
        content_frame.grid_rowconfigure(1, weight=0)  
        
        self.listas_frame = ctk.CTkFrame(content_frame, fg_color="white")
        self.listas_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        self.listas_frame.grid_rowconfigure(0, weight=1)
        self.listas_frame.grid_columnconfigure(0, weight=1)

        botoes_frame = ctk.CTkFrame(content_frame, fg_color="white")
        botoes_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 10))
        botoes_frame.grid_columnconfigure(0, weight=1)
        botoes_frame.grid_columnconfigure(1, weight=0)
        botoes_frame.grid_columnconfigure(2, weight=0)
        botoes_frame.grid_columnconfigure(3, weight=0)
        botoes_frame.grid_columnconfigure(4, weight=1)

        botoes_acoes = [
            ("➕Cadastrar", "#4CAF50", self.abrir_cadastro_menu),
            ("✏️Editar", "#047194", self.abrir_editar),
            ("🗑️Excluir", "#e74c3c", self.abrir_excluir)
        ]

        for i, (texto, cor, comando) in enumerate(botoes_acoes):
            borda_botoes = {
                "➕Cadastrar": "#2E7D32",
                "✏️Editar": "#034b66",
                "🗑️Excluir": "#c0392b"
            }[texto]

            hover_botoes = {
                "➕Cadastrar": "#66BB6A",
                "✏️Editar": "#0490c7",
                "🗑️Excluir": "#ff6b5c"
            }[texto]

            btn = ctk.CTkButton(
                botoes_frame,
                text=texto,
                command=comando,
                fg_color=cor,
                hover_color=hover_botoes,
                text_color="#fff",
                font=("Segoe UI Emoji", 14, "bold"),
                height=35,
                width=100,
                border_width=2,
                border_color=borda_botoes
            )
            btn.grid(row=0, column=i+1, padx=5, pady=5)
            if texto == "✏️Editar":
                self.btn_editar = btn
            elif texto == "🗑️Excluir":
                self.btn_excluir = btn

        self.btn_editar.configure(state="disabled")
        self.btn_excluir.configure(state="disabled")

    def construir_abas(self):
        # Se o notebook já existe, não recriar
        if hasattr(self, 'notebook'):
            return

        # Criar e configurar o estilo ANTES de criar o notebook
        style = ttk.Style()
        
        # Configurar o estilo principal do notebook
        style.configure("TNotebook", 
                    background="#ffffff",
                    borderwidth=0,
                    tabmargins=[0, 0, 0, 2])
        
        # Configurar o estilo das abas - ABA SELECIONADA MAIOR
        style.configure("TNotebook.Tab",
                    padding=[13, 4],  # Padding MENOR para abas não selecionadas
                    font=('Roboto', 12, 'bold'),
                    background="#f0f0f0",
                    foreground="#333333",
                    focuscolor=style.lookup("TNotebook.Tab", "background"))
        
        # Mapear estados - ABA SELECIONADA COM MAIOR PADDING
        style.map("TNotebook.Tab",
                background=[("selected", "#ffffff"),
                            ("active", "#e6e6e6")],
                foreground=[("selected", "black"),
                            ("active", "#333333")],
                padding=[("selected", [18, 6]),  # Padding MAIOR quando selecionada
                        ("active", [13, 4]),
                        ("!selected", [13, 4])])  # Padding menor quando não selecionada

        self.notebook = ttk.Notebook(self.listas_frame, style="TNotebook")
        self.notebook.pack(fill="both", expand=True, padx=5, pady=5)

        self.treeviews = {}

        tipos = ["cliente", "funcionario", "fornecedor"]
        titulos = ["Clientes", "Funcionários", "Fornecedores"]

        for tipo, titulo in zip(tipos, titulos):
            frame_aba = ctk.CTkFrame(self.notebook, fg_color="white")
            self.notebook.add(frame_aba, text=titulo)

            container = ctk.CTkFrame(frame_aba, fg_color="white")
            container.pack(fill="both", expand=True, padx=10, pady=10)
            container.grid_rowconfigure(0, weight=1)
            container.grid_columnconfigure(0, weight=1)

            if tipo == "cliente":
                colunas = ["Nome", "CPF", "Data Nasc.", "Idade", "Email", "CEP", "Endereço", "Número", "Bairro", "Cidade", "Estado"]
            elif tipo == "funcionario":
                colunas = ["Nome", "CPF", "Data Nasc.", "Idade", "Departamento", "Setor", "Data Admissão", "Email"]
            else:
                colunas = ["Nome", "CNPJ", "Tipo Fornecimento", "Email", "CEP", "Endereço", "Número", "Bairro", "Cidade", "Estado"]

            # Configurar estilo da Treeview
            tree_style = ttk.Style()
            tree_style.theme_use("clam")
            tree_style.configure("Treeview.Heading", font=("Roboto", 10, "bold"), background="#ebebeb")
            tree_style.configure("Treeview", 
                                font=("Segoe UI", 10), 
                                background="white", 
                                fieldbackground="white", 
                                foreground="black",
                                rowheight=25)
            
            tree_style.map("Treeview",
                          background=[("selected", "#e6f2ff")],
                          foreground=[("selected", "black")])

            treeview = ttk.Treeview(container, columns=colunas, show="headings", height=15, style="Treeview")
            
            for col in colunas:
                treeview.heading(col, text=col)
                treeview.column(col, width=100, anchor="w")

            if tipo == "cliente":
                treeview.column("Nome", width=150)
                treeview.column("Email", width=150)
                treeview.column("Endereço", width=150)
            elif tipo == "funcionario":
                treeview.column("Nome", width=150)
                treeview.column("Departamento", width=120)
                treeview.column("Setor", width=120)
            else:
                treeview.column("Nome", width=150)
                treeview.column("Tipo Fornecimento", width=150)
                treeview.column("Email", width=150)

            vsb = ttk.Scrollbar(container, orient="vertical", command=treeview.yview)
            treeview.configure(yscrollcommand=vsb.set)
            
            hsb = ttk.Scrollbar(container, orient="horizontal", command=treeview.xview)
            treeview.configure(xscrollcommand=hsb.set)

            treeview.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            hsb.grid(row=1, column=0, sticky="ew")

            self.treeviews[tipo] = treeview

            treeview.bind('<<TreeviewSelect>>', lambda e, t=tipo: self.on_treeview_select(e, t))
            treeview.bind('<Double-1>', lambda e, t=tipo: self.on_double_click(e, t))

            treeview.tag_configure("selected", background="#e6f2ff", foreground="black")

        # Iniciar a verificação periódica do estilo
        self.aplicar_estilo_persistente()

    def aplicar_estilo_persistente(self):
        """Aplicar e manter o estilo das abas periodicamente para evitar reversão"""
        try:
            # Re-aplicar o estilo a cada 3 segundos
            style = ttk.Style()
            style.configure("TNotebook.Tab",
                           padding=[15, 6],  # Padding padrão menor
                           font=('Roboto', 14, 'bold'),
                           background="#f0f0f0",
                           foreground="#333333")
            
            # Mapear para aba selecionada ter padding maior
            style.map("TNotebook.Tab",
                     background=[("selected", "#ffffff"),
                                ("active", "#e6e6e6")],
                     foreground=[("selected", "black"),
                                ("active", "#333333")],
                     padding=[("selected", [20, 8]),  # MAIOR quando selecionada
                             ("active", [15, 6]),
                             ("!selected", [15, 6])])  # MENOR quando não selecionada
        except Exception as e:
            print(f"Estilo não pôde ser reaplicado: {e}")
        
        # Agendar próxima verificação
        self.after(3000, self.aplicar_estilo_persistente)

    def menu_import_export_selecionado(self, opcao):
        if opcao == "Importar":
            self.abrir_importar()
        elif opcao == "Exportar":
            self.abrir_exportar()

    def abrir_agenda(self):
        try:
            if os.path.exists("COD_AGENDA.py"):
                subprocess.Popen([sys.executable, "COD_AGENDA.py"])
            else:
                messagebox.showwarning("Aviso", "Arquivo da agenda não encontrado!")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível abrir a agenda:\n{str(e)}")

    def menu_selecionado(self, opcao):
        if opcao == "⚙️ Configurações":
            self.alterar_informacoes_usuario()

    def atualizar_relogio(self):
        now = time.localtime()
        date_str = time.strftime("%d/%m/%Y", now)
        time_str = time.strftime("%H:%M:%S", now)

        self.date_label.configure(text=date_str)
        self.time_label.configure(text=time_str)
        self.after(1000, self.atualizar_relogio)

    def sair(self):
        self.destroy()
        self.master.deiconify()

    def abrir_ajuda(self):
        try:
            if os.path.exists("COD_GUIDE.py"):
                subprocess.Popen([sys.executable, "COD_GUIDE.py"])
            else:
                messagebox.showwarning("Aviso", "Arquivo de ajuda não encontrado!")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível abrir o guia:\n{str(e)}")

    def abrir_sobre(self):
        try:
            if os.path.exists("COD_ON.py"):
                subprocess.Popen([sys.executable, "COD_ON.py"])
            else:
                messagebox.showwarning("Aviso", "Arquivo de sobre não encontrado!")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível abrir a tela sobre:\n{str(e)}")
    
    def abrir_importar(self):
        win = TelaImportar(self)
        win.grab_set()
    
    def abrir_exportar(self):
        win = TelaExportar(self)
        win.grab_set()
    
    def abrir_cadastro_menu(self):
        win = TelaCadastroPessoa(self)
        win.grab_set()
        win.wait_window()
        self.atualizar_lista()

    def abrir_editar(self):
        if not hasattr(self, 'treeview_selecionada') or not self.treeview_selecionada.selection():
            messagebox.showwarning("Aviso", "Selecione um registro para editar.")
            return

        selected_item = self.treeview_selecionada.selection()[0]
        valores = self.treeview_selecionada.item(selected_item, 'values')
        tipo = self.tipo_selecionado

        if tipo == "fornecedor":
            documento = valores[1]
        else:
            documento = valores[1]

        pessoa, _ = buscar_pessoa_por_documento(documento)
        if pessoa:
            win = TelaCadastroPessoa(self, pessoa)
            win.grab_set()
            win.wait_window()
            self.atualizar_lista()
        else:
            messagebox.showerror("Erro", "Pessoa não encontrada.")

    def abrir_excluir(self):
        if not hasattr(self, 'treeview_selecionada') or not self.treeview_selecionada.selection():
            messagebox.showwarning("Aviso", "Selecione um registro para excluir.")
            return

        selected_item = self.treeview_selecionada.selection()[0]
        valores = self.treeview_selecionada.item(selected_item, 'values')
        tipo = self.tipo_selecionado

        if tipo == "fornecedor":
            documento = valores[1]
        else:
            documento = valores[1]

        pessoa, _ = buscar_pessoa_por_documento(documento)
        if pessoa:
            resposta = messagebox.askyesno(
                "Confirmar Exclusão", 
                f"Tem certeza que deseja excluir {tratar_valor_vazio(pessoa['nome'])}?"
            )
            if resposta:
                arquivo = ARQUIVO_PESSOAS[tipo]
                pessoas = carregar_dados_banco(arquivo)
                pessoas = [p for p in pessoas if p.get("cpf") != pessoa.get("cpf") and p.get("cnpj") != pessoa.get("cnpj")]
                salvar_dados_banco(arquivo, pessoas)
                messagebox.showinfo("Sucesso", "Pessoa excluída com sucesso!")
                self.atualizar_lista()
        else:
            messagebox.showerror("Erro", "Pessoa não encontrada.")

    def alterar_informacoes_usuario(self):
        win = TelaRegistroUsuario(self, self.user)
        win.grab_set()
        win.wait_window()
        self.atualizar_info_display_usuario()

    def atualizar_info_display_usuario(self):
        users = carregar_dados_banco(ARQUIVO_USUARIOS)
        dados_corretos_usuario = next((u for u in users if u["usuario"] == self.user["usuario"]), None)
        if dados_corretos_usuario:
            self.user = dados_corretos_usuario
            for widget in self.grid_slaves():
                if widget.grid_info()["row"] == 0:
                    widget.destroy()
            self.criar_header()
        else:
            messagebox.showerror("Erro", "Usuário atual não encontrado após alteração. Por favor, faça login novamente.")
            self.master.deiconify()
            self.destroy()

    def atualizar_lista(self):
        # Se as treeviews ainda não foram criadas, não fazer nada
        if not hasattr(self, 'treeviews'):
            return

        for tipo, treeview in self.treeviews.items():
            self.preencher_treeview(tipo, treeview)

    def preencher_treeview(self, tipo, treeview):
        for item in treeview.get_children():
            treeview.delete(item)
        
        arquivo = ARQUIVO_PESSOAS[tipo]
        pessoas = carregar_dados_banco(arquivo)
        
        for pessoa in pessoas:
            if tipo == "cliente":
                valores = [
                    tratar_valor_vazio(pessoa.get("nome", "")),
                    tratar_valor_vazio(pessoa.get("cpf", "")),
                    tratar_valor_vazio(pessoa.get("data_nasc", "")),
                    tratar_valor_vazio(pessoa.get("idade", "")),
                    tratar_valor_vazio(pessoa.get("email", "")),
                    tratar_valor_vazio(pessoa.get("cep", "")),
                    tratar_valor_vazio(pessoa.get("endereco", "")),
                    tratar_valor_vazio(pessoa.get("numero", "")),
                    tratar_valor_vazio(pessoa.get("bairro", "")),
                    tratar_valor_vazio(pessoa.get("cidade", "")),
                    tratar_valor_vazio(pessoa.get("estado", ""))
                ]
            elif tipo == "funcionario":
                valores = [
                    tratar_valor_vazio(pessoa.get("nome", "")),
                    tratar_valor_vazio(pessoa.get("cpf", "")),
                    tratar_valor_vazio(pessoa.get("data_nasc", "")),
                    tratar_valor_vazio(pessoa.get("idade", "")),
                    tratar_valor_vazio(pessoa.get("departamento", "")),
                    tratar_valor_vazio(pessoa.get("setor", "")),
                    tratar_valor_vazio(pessoa.get("data_admissao", "")),
                    tratar_valor_vazio(pessoa.get("email", ""))
                ]
            else:
                valores = [
                    tratar_valor_vazio(pessoa.get("nome", "")),
                    tratar_valor_vazio(pessoa.get("cnpj", "")),
                    tratar_valor_vazio(pessoa.get("tipo_fornecimento", "")),
                    tratar_valor_vazio(pessoa.get("email", "")),
                    tratar_valor_vazio(pessoa.get("cep", "")),
                    tratar_valor_vazio(pessoa.get("endereco", "")),
                    tratar_valor_vazio(pessoa.get("numero", "")),
                    tratar_valor_vazio(pessoa.get("bairro", "")),
                    tratar_valor_vazio(pessoa.get("cidade", "")),
                    tratar_valor_vazio(pessoa.get("estado", ""))
                ]
            
            treeview.insert("", "end", values=valores)

    def on_treeview_select(self, event, tipo):
        treeview = self.treeviews[tipo]
        selected = treeview.selection()
        if selected:
            self.btn_editar.configure(state="normal")
            self.btn_excluir.configure(state="normal")
            self.treeview_selecionada = treeview
            self.tipo_selecionado = tipo
        else:
            self.btn_editar.configure(state="disabled")
            self.btn_excluir.configure(state="disabled")

    def on_double_click(self, event, tipo):
        treeview = self.treeviews[tipo]
        selected = treeview.selection()
        if selected:
            self.abrir_editar()

if __name__ == "__main__":
    os.makedirs("fotos", exist_ok=True)
    if not os.path.exists(SEM_IMAGEM):
        try:
            placeholder_img = Image.new('RGB', (50, 50), color = 'black')
            placeholder_img.save(SEM_IMAGEM)
        except Exception as e:
            print(f"Erro ao criar imagem padrão: {e}")

    for arquivo in ARQUIVO_PESSOAS.values():
        if not os.path.exists(arquivo):
            with open(arquivo, "w") as f:
                json.dump([], f)

    ctk.set_appearance_mode("light")
    app = TelaLogin()
    app.mainloop()